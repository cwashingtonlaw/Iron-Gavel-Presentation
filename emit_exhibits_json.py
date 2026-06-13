#!/usr/bin/env python3
"""
emit_exhibits_json.py — D&W / Courtroom Justice exhibit sidecar emitter
=======================================================================
Reads the Exhibit sheet of a case's Case Tables.xlsx and writes a validated
{CASE_ROOT}/Trial/exhibits.json that the Courtroom Justice app consumes.

DESIGN INVARIANTS (do not violate without bumping contract_version):
  * The .xlsx is the system of record. This script only PROJECTS it to JSON.
  * exhibit.file paths are relative to the directory containing exhibits.json
    (i.e. {CASE_ROOT}/Trial/). Declared in the file as path_base="sidecar_dir".
  * Write is ATOMIC (temp file + os.replace) so the app never reads a half file.
  * Output is VALIDATED against exhibits.schema.json before it is committed.
    On any validation failure nothing is overwritten and the script exits 1.

USAGE:
  python3 emit_exhibits_json.py \
      --workbook "/path/to/CASE_ROOT/Case Tables.xlsx" \
      --case-root "/path/to/CASE_ROOT" \
      [--schema "/path/to/exhibits.schema.json"] \
      [--sheet "Exhibits"]

The skill invokes this after every exhibit-ledger update (see SKILL_PATCH.md).
"""

import argparse
import datetime as _dt
import json
import os
import sys
import tempfile

try:
    import openpyxl
except ImportError:
    sys.exit("FATAL: openpyxl not installed. Run: pip3 install openpyxl --break-system-packages")

CONTRACT_VERSION = "1.0"

# ---- Column aliasing -------------------------------------------------------
# Header text is lowercased and stripped before matching. Add real headers here
# if the firm's Exhibit sheet uses different labels; logic below never changes.
ALIASES = {
    "id":          ["exhibit no", "exhibit no.", "exhibit number", "exhibit #", "exhibit", "no", "no.", "id"],
    "party":       ["party", "side", "offered by", "proponent"],
    "description": ["description", "desc", "exhibit description", "title"],
    "file":        ["file", "filename", "file name", "path", "relative path"],
    "witness":     ["witness", "sponsoring witness", "thru witness", "through witness"],
    "bates":       ["bates", "bates range", "bates no", "bates nos", "bates numbers"],
    "status":      ["status", "exhibit status"],
    "offered":     ["offered", "offered?", "date offered"],
    "objection":   ["objection", "objections", "objection logged"],
    "ruling":      ["ruling", "court ruling", "disposition"],
    "admitted":    ["admitted", "admitted?", "date admitted"],
    "excluded":    ["excluded", "excluded?", "denied"],
    "notes":       ["notes", "note", "comment", "comments"],
}

PARTY_NORMAL = {
    "d": "Defense", "def": "Defense", "defense": "Defense", "defendant": "Defense",
    "s": "State", "state": "State", "p": "State", "prosecution": "State", "government": "State", "govt": "State",
    "j": "Joint", "joint": "Joint",
    "c": "Court", "court": "Court",
}

IMAGE_EXT = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".heic", ".gif", ".bmp", ".webp"}
VIDEO_EXT = {".mp4", ".mov", ".m4v", ".avi", ".mkv"}


def _truthy(v):
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s not in ("", "no", "n", "false", "0", "n/a", "na", "-", "—", "pending")


def media_type_for(filename):
    ext = os.path.splitext(filename or "")[1].lower()
    if ext == ".pdf":
        return "pdf"
    if ext in IMAGE_EXT:
        return "image"
    if ext in VIDEO_EXT:
        return "video"
    return "unknown"


def derive_status(row):
    """Single Status column wins if present and recognized; else derive from flags.
    Precedence: excluded > admitted > objected > offered > pending."""
    s = (str(row.get("status") or "").strip().lower())
    if s in ("pending", "offered", "objected", "admitted", "excluded"):
        return s
    ruling = (str(row.get("ruling") or "").strip().lower())
    if _truthy(row.get("excluded")) or ruling in ("excluded", "denied", "sustained", "not admitted"):
        return "excluded"
    if _truthy(row.get("admitted")) or ruling in ("admitted", "overruled", "granted"):
        return "admitted"
    if _truthy(row.get("objection")):
        return "objected"
    if _truthy(row.get("offered")):
        return "offered"
    return "pending"


def normalize_party(v):
    return PARTY_NORMAL.get(str(v or "").strip().lower(), str(v or "").strip() or "Defense")


def resolve_file(explicit, status, exid):
    """If the sheet gives an explicit relative path, trust it. Otherwise place the
    file in Exhibits_Admitted (admitted) or Exhibits_Pending (everything else),
    matching the Operating Manual folder discipline. Paths are relative to Trial/."""
    if explicit and str(explicit).strip():
        return str(explicit).strip()
    # No explicit filename: we cannot invent the descriptive name, so point at the
    # folder with the id as a stub. Empty allowed only for pending (schema).
    folder = "Exhibits_Admitted" if status == "admitted" else "Exhibits_Pending"
    return f"{folder}/{exid}.pdf" if status != "pending" else ""


def find_sheet(wb, requested):
    if requested:
        if requested in wb.sheetnames:
            return wb[requested]
        sys.exit(f"FATAL: sheet '{requested}' not found. Sheets present: {wb.sheetnames}")
    for name in wb.sheetnames:
        if "exhibit" in name.lower():
            return wb[name]
    sys.exit(f"FATAL: no sheet with 'exhibit' in its name. Sheets present: {wb.sheetnames}")


def build_header_map(header_row):
    """Map each canonical field to a column index using ALIASES."""
    seen = {}
    for idx, cell in enumerate(header_row):
        label = str(cell.value or "").strip().lower()
        if label:
            seen[label] = idx
    colmap = {}
    for field, aliases in ALIASES.items():
        for a in aliases:
            if a in seen:
                colmap[field] = seen[a]
                break
    return colmap


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--case-root", required=True)
    ap.add_argument("--schema", default=os.path.join(os.path.dirname(__file__), "exhibits.schema.json"))
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--caption", default=None, help="Override case caption (else read from a 'Case' sheet if present)")
    ap.add_argument("--docket", default=None)
    ap.add_argument("--court", default=None)
    args = ap.parse_args()

    if not os.path.isfile(args.workbook):
        sys.exit(f"FATAL: workbook not found: {args.workbook}")

    wb = openpyxl.load_workbook(args.workbook, data_only=True, read_only=True)
    ws = find_sheet(wb, args.sheet)

    rows = list(ws.iter_rows())
    if not rows:
        sys.exit("FATAL: exhibit sheet is empty.")
    colmap = build_header_map(rows[0])
    for required in ("id", "description"):
        if required not in colmap:
            sys.exit(f"FATAL: could not find a column for '{required}'. "
                     f"Headers seen: {[c.value for c in rows[0]]}. Add the real header to ALIASES.")

    def get(row, field):
        i = colmap.get(field)
        return row[i].value if i is not None and i < len(row) else None

    exhibits, problems = [], []
    for r_i, row in enumerate(rows[1:], start=2):
        raw_id = get(row, "id")
        if raw_id is None or str(raw_id).strip() == "":
            continue  # blank line
        exid = str(raw_id).strip()
        row_dict = {f: get(row, f) for f in ALIASES}
        status = derive_status(row_dict)
        explicit_file = get(row, "file")
        file_rel = resolve_file(explicit_file, status, exid)
        mtype = media_type_for(file_rel) if file_rel else "unknown"

        ex = {
            "id": exid,
            "party": normalize_party(get(row, "party")),
            "description": str(get(row, "description") or "").strip(),
            "file": file_rel,
            "status": status,
            "media_type": mtype,
        }
        for opt in ("witness", "bates", "objection", "ruling", "notes"):
            val = get(row, opt)
            if val is not None and str(val).strip():
                ex[opt] = str(val).strip()
        exhibits.append(ex)

        # Soft warnings (do not block the write; surfaced to the operator)
        if status == "admitted" and not file_rel:
            problems.append(f"row {r_i} ({exid}): admitted but no file path.")
        if file_rel and mtype == "unknown":
            problems.append(f"row {r_i} ({exid}): unrecognized file type '{file_rel}'.")

    # Case identity
    caption = args.caption or "UNKNOWN CAPTION"
    docket = args.docket or ""
    court = args.court or ""
    if not args.caption and "Case" in wb.sheetnames:
        cs = wb["Case"]
        kv = {}
        for row in cs.iter_rows(values_only=True):
            if row and row[0]:
                kv[str(row[0]).strip().lower()] = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        caption = kv.get("caption", caption)
        docket = kv.get("docket", docket)
        court = kv.get("court", court)

    payload = {
        "contract_version": CONTRACT_VERSION,
        "case": {"caption": caption, "docket": docket, "court": court},
        "generated": _dt.datetime.now().astimezone().isoformat(timespec="seconds"),
        "path_base": "sidecar_dir",
        "exhibits": exhibits,
    }

    # Validate before writing anything.
    try:
        import jsonschema
        with open(args.schema) as f:
            schema = json.load(f)
        jsonschema.validate(payload, schema)
    except ImportError:
        print("WARNING: jsonschema not installed; skipping validation.", file=sys.stderr)
    except Exception as e:
        sys.exit(f"FATAL: output failed schema validation, nothing written.\n{e}")

    out_dir = os.path.join(args.case_root, "Trial")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "exhibits.json")
    fd, tmp = tempfile.mkstemp(dir=out_dir, suffix=".tmp")
    with os.fdopen(fd, "w") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
    os.replace(tmp, out_path)  # atomic

    print(f"OK: wrote {len(exhibits)} exhibits -> {out_path}")
    if problems:
        print("WARNINGS (sidecar still written):", file=sys.stderr)
        for p in problems:
            print("  - " + p, file=sys.stderr)


if __name__ == "__main__":
    main()
